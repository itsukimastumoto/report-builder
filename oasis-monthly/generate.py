#!/usr/bin/env python3
"""
月次実績集計レポート生成ツール

input/フォルダのExcelファイルから、キャンペーン販促費・ポイント利用額のレポートを自動生成する。

使い方:
    # input/フォルダにExcelファイルを配置して実行
    ./generate.py 202601

    # ファイルを直接指定する場合
    ./generate.py 202601 --campaign file1.xlsx --discount file2.xlsx --point file3.xlsx
"""

import argparse
import re
import shutil
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side


SCRIPT_DIR = Path(__file__).parent
TEMPLATE_DIR = SCRIPT_DIR / "templates"
INPUT_DIR = SCRIPT_DIR / "input"
OUTPUT_DIR = SCRIPT_DIR / "output"

CAMPAIGN_TEMPLATE = TEMPLATE_DIR / "yyyymm_【月次集計】キャンペーン販促費.xlsx"
POINT_TEMPLATE = TEMPLATE_DIR / "yyyymm_【月次集計】ポイント利用額.xlsx"


def get_value(row, col_name, default=None):
    """pandasの行から値を取得（NaN対応）"""
    if col_name not in row:
        return default
    val = row[col_name]
    if pd.isna(val):
        return default
    return val


def read_data(file_path: Path) -> pd.DataFrame:
    """CSVまたはExcelファイルを読み込む"""
    suffix = file_path.suffix.lower()
    if suffix == '.csv':
        return pd.read_csv(file_path)
    elif suffix in ['.xlsx', '.xls']:
        return pd.read_excel(file_path)
    else:
        raise ValueError(f"未対応のファイル形式: {suffix}")


def find_input_files(input_dir: Path) -> dict:
    """input/フォルダからファイルを自動検出"""
    files = {
        'campaign': None,
        'discount': None,
        'point': None,
    }

    for f in input_dir.glob('*'):
        if f.suffix.lower() not in ['.csv', '.xlsx', '.xls']:
            continue
        name_lower = f.name.lower()

        if '商品×割引率' in f.name or '商品×割引率' in f.name or 'campaign' in name_lower:
            files['campaign'] = f
        elif '割引率一覧' in f.name or 'discount' in name_lower:
            files['discount'] = f
        elif 'ポイント利用' in f.name or 'point' in name_lower:
            files['point'] = f

    return files


def generate_campaign_report(month: str, campaign_file: Path, discount_file: Path, output_dir: Path) -> Path:
    """キャンペーン販促費レポートを生成"""

    df_campaign = read_data(campaign_file)
    df_discount = read_data(discount_file)

    output_file = output_dir / f"{month}_【月次集計】キャンペーン販促費.xlsx"
    shutil.copy(CAMPAIGN_TEMPLATE, output_file)

    wb = load_workbook(output_file)

    # 元データシート1: キャンペーン販売商品とその割引総額
    ws_campaign = wb["(元データ)キャンペーン販売商品とその割引総額"]
    for row in range(2, ws_campaign.max_row + 1):
        for col in range(1, 12):
            ws_campaign.cell(row=row, column=col).value = None

    for idx, row in df_campaign.iterrows():
        excel_row = idx + 2
        ws_campaign.cell(row=excel_row, column=1).value = get_value(row, 'brand_name', '')
        ws_campaign.cell(row=excel_row, column=2).value = get_value(row, 'item_name', '')
        ws_campaign.cell(row=excel_row, column=3).value = get_value(row, 'item_code', '')
        ws_campaign.cell(row=excel_row, column=4).value = get_value(row, 'discount_rate', 0)
        ws_campaign.cell(row=excel_row, column=5).value = get_value(row, 'issue_count', 0)
        ws_campaign.cell(row=excel_row, column=6).value = get_value(row, 'item_price', 0)
        ws_campaign.cell(row=excel_row, column=7).value = get_value(row, 'total_item_price', 0)
        ws_campaign.cell(row=excel_row, column=8).value = get_value(row, 'purchase_price', 0)
        ws_campaign.cell(row=excel_row, column=9).value = get_value(row, 'total_purchase_amount', 0)
        ws_campaign.cell(row=excel_row, column=10).value = get_value(row, 'discount_amount', 0)
        ws_campaign.cell(row=excel_row, column=11).value = get_value(row, 'total_discount_amount', 0)

    campaign_data_rows = len(df_campaign)

    # 元データシート2: 基本割引率一覧
    ws_discount = wb["(元データ)基本割引率一覧"]
    for row in range(2, ws_discount.max_row + 1):
        for col in range(1, 15):
            ws_discount.cell(row=row, column=col).value = None

    discount_columns = ['ブランド', '商品', 'gx_item_type', 'gx_item_code', '等価価格', '割引率',
                        '割引価格', 'おすすめ度', '掲載状況', '案件個別_掲載可', 'マスタ_掲載可',
                        '更新日', 'project_id', 'client_id']

    for idx, row in df_discount.iterrows():
        excel_row = idx + 2
        for col_idx, col_name in enumerate(discount_columns, 1):
            val = get_value(row, col_name)
            if val is not None:
                ws_discount.cell(row=excel_row, column=col_idx).value = val

    discount_data_rows = len(df_discount)

    # キャンペーン費用集計シートを更新
    ws_calc = wb["キャンペーン費用集計"]
    for row in range(4, ws_calc.max_row + 1):
        for col in range(1, 21):
            ws_calc.cell(row=row, column=col).value = None

    for idx, row in df_campaign.iterrows():
        excel_row = idx + 4
        ws_calc.cell(row=excel_row, column=1).value = get_value(row, 'brand_name', '')
        ws_calc.cell(row=excel_row, column=2).value = get_value(row, 'item_name', '')
        ws_calc.cell(row=excel_row, column=3).value = get_value(row, 'item_code', '')
        ws_calc.cell(row=excel_row, column=4).value = get_value(row, 'discount_rate', 0)
        ws_calc.cell(row=excel_row, column=5).value = get_value(row, 'issue_count', 0)
        ws_calc.cell(row=excel_row, column=6).value = get_value(row, 'item_price', 0)
        ws_calc.cell(row=excel_row, column=7).value = get_value(row, 'total_item_price', 0)
        ws_calc.cell(row=excel_row, column=8).value = get_value(row, 'purchase_price', 0)
        ws_calc.cell(row=excel_row, column=9).value = get_value(row, 'total_purchase_amount', 0)
        ws_calc.cell(row=excel_row, column=10).value = get_value(row, 'discount_amount', 0)
        ws_calc.cell(row=excel_row, column=11).value = get_value(row, 'total_discount_amount', 0)

        # 数式を設定
        ws_calc.cell(row=excel_row, column=12).value = \
            f"=VLOOKUP(キャンペーン費用集計!$C{excel_row},'(元データ)基本割引率一覧'!D$2:L${discount_data_rows + 1},3,)"
        ws_calc.cell(row=excel_row, column=13).value = \
            f"=VLOOKUP(キャンペーン費用集計!$C{excel_row},'(元データ)基本割引率一覧'!D$2:N${discount_data_rows + 1},4,)"
        ws_calc.cell(row=excel_row, column=14).value = f"=M{excel_row}*E{excel_row}"
        ws_calc.cell(row=excel_row, column=15).value = f"=ROUNDDOWN(F{excel_row}*L{excel_row},0)"
        ws_calc.cell(row=excel_row, column=16).value = f"=O{excel_row}*E{excel_row}"
        ws_calc.cell(row=excel_row, column=17).value = f"=G{excel_row}-N{excel_row}"
        ws_calc.cell(row=excel_row, column=18).value = f'=IF(Q{excel_row}=P{excel_row},"OK","")'
        ws_calc.cell(row=excel_row, column=19).value = f"=K{excel_row}-P{excel_row}"
        # キャンペーン判定: 割引率(D列)と基本割引率(L列)を比較
        ws_calc.cell(row=excel_row, column=20).value = f'=IF($D{excel_row}=$L{excel_row},"通常","キャンペーン")'

    last_data_row = campaign_data_rows + 3
    ws_calc.cell(row=3, column=11).value = f"=SUM(K4:K{last_data_row})"
    ws_calc.cell(row=3, column=16).value = f"=SUM(P4:P{last_data_row})"
    ws_calc.cell(row=3, column=17).value = f"=SUM(Q4:Q{last_data_row})"
    ws_calc.cell(row=3, column=19).value = f"=SUM(S4:S{last_data_row})"

    wb.save(output_file)
    return output_file


def generate_point_report(month: str, point_file: Path, output_dir: Path) -> Path:
    """ポイント利用額レポートを生成"""

    df_point = read_data(point_file)

    output_file = output_dir / f"{month}_【月次集計】ポイント利用額.xlsx"
    shutil.copy(POINT_TEMPLATE, output_file)

    wb = load_workbook(output_file)

    # 元データシート
    ws_raw = wb["(元データ) oasis ポイント利用状況(月次)"]
    for row in range(2, ws_raw.max_row + 1):
        for col in range(1, 11):
            ws_raw.cell(row=row, column=col).value = None

    raw_columns = ['client_id', 'company_name', 'month', 'granted_points', 'used_points',
                   'remaining_points', 'avg_user_usage', 'point_usage_rate_percent',
                   'monthly_point_users', 'point_consumption_rate_percent']

    for idx, row in df_point.iterrows():
        excel_row = idx + 2
        for col_idx, col_name in enumerate(raw_columns, 1):
            val = get_value(row, col_name)
            if val is not None:
                ws_raw.cell(row=excel_row, column=col_idx).value = val

    # 集計結果シート
    ws_result = wb["集計結果"]
    # 既存データをクリア（ヘッダー以外全て）
    for row in range(2, ws_result.max_row + 1):
        for col in range(1, 6):
            ws_result.cell(row=row, column=col).value = None

    result_columns = ['client_id', 'company_name', 'month', 'granted_points', 'used_points']

    for idx, row in df_point.iterrows():
        excel_row = idx + 2
        for col_idx, col_name in enumerate(result_columns, 1):
            val = get_value(row, col_name, 0 if col_name in ['granted_points', 'used_points'] else '')
            ws_result.cell(row=excel_row, column=col_idx).value = val

    # 合計行をデータ直後に配置
    last_data_row = len(df_point) + 1
    sum_row = last_data_row + 1
    ws_result.cell(row=sum_row, column=2).value = "合計利用pt"
    sum_cell = ws_result.cell(row=sum_row, column=5)
    sum_cell.value = f"=SUM(E2:E{last_data_row})"
    sum_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 枠線スタイル
    medium = Side(style="medium")
    no_border = Side(style=None)

    # 合計行に太枠（下端）を設定
    ws_result.cell(row=sum_row, column=1).border = Border(left=medium, bottom=medium)
    ws_result.cell(row=sum_row, column=2).border = Border(bottom=medium)
    ws_result.cell(row=sum_row, column=3).border = Border(bottom=medium)
    ws_result.cell(row=sum_row, column=4).border = Border(bottom=medium)
    ws_result.cell(row=sum_row, column=5).border = Border(right=medium, bottom=medium)

    # テンプレート28行目の注記（F28）をクリア
    ws_result.cell(row=28, column=6).value = None

    # 28行目が合計行でない場合、28行目以降の枠線と値をクリア
    if sum_row != 28:
        for row in range(sum_row + 1, 29):
            for col in range(1, 6):
                ws_result.cell(row=row, column=col).value = None
                ws_result.cell(row=row, column=col).fill = PatternFill(fill_type=None)
                ws_result.cell(row=row, column=col).border = Border()

    wb.save(output_file)
    return output_file


def main():
    parser = argparse.ArgumentParser(description="月次実績集計レポート生成ツール")
    parser.add_argument("month", help="対象月 (YYYYMM形式)")
    parser.add_argument("--campaign", type=Path, help="キャンペーン販売商品データ")
    parser.add_argument("--discount", type=Path, help="基本割引率一覧")
    parser.add_argument("--point", type=Path, help="ポイント利用状況")
    parser.add_argument("--output", type=Path, default=None, help="出力先ディレクトリ")

    args = parser.parse_args()

    # input/フォルダから自動検出（引数で指定されていない場合）
    detected = find_input_files(INPUT_DIR)

    campaign_file = args.campaign or detected['campaign']
    discount_file = args.discount or detected['discount']
    point_file = args.point or detected['point']

    output_dir = args.output or OUTPUT_DIR / args.month
    output_dir.mkdir(parents=True, exist_ok=True)

    results = []

    # キャンペーン販促費
    if campaign_file and discount_file:
        if not campaign_file.exists():
            print(f"エラー: ファイルが見つかりません: {campaign_file}")
            sys.exit(1)
        if not discount_file.exists():
            print(f"エラー: ファイルが見つかりません: {discount_file}")
            sys.exit(1)

        print(f"キャンペーンデータ: {campaign_file.name}")
        print(f"割引率データ: {discount_file.name}")

        output_file = generate_campaign_report(args.month, campaign_file, discount_file, output_dir)
        results.append(f"キャンペーン販促費: {output_file}")
    else:
        missing = []
        if not campaign_file:
            missing.append("キャンペーン販売商品データ")
        if not discount_file:
            missing.append("基本割引率一覧")
        print(f"スキップ: キャンペーン販促費（{', '.join(missing)}が見つかりません）")

    # ポイント利用額
    if point_file:
        if not point_file.exists():
            print(f"エラー: ファイルが見つかりません: {point_file}")
            sys.exit(1)

        print(f"ポイントデータ: {point_file.name}")

        output_file = generate_point_report(args.month, point_file, output_dir)
        results.append(f"ポイント利用額: {output_file}")
    else:
        print("スキップ: ポイント利用額（データが見つかりません）")

    if results:
        print("\n生成完了:")
        for r in results:
            print(f"  {r}")

        # inputファイルをoutput/YYYYMM/input/にバックアップとして移動
        input_backup_dir = output_dir / "input"
        input_backup_dir.mkdir(exist_ok=True)

        moved_files = []
        for key, filepath in [('campaign', campaign_file), ('discount', discount_file), ('point', point_file)]:
            if filepath and filepath.exists() and filepath.parent == INPUT_DIR:
                dest = input_backup_dir / filepath.name
                shutil.move(str(filepath), str(dest))
                moved_files.append(filepath.name)

        if moved_files:
            print(f"\ninputファイルを {input_backup_dir} に移動しました:")
            for f in moved_files:
                print(f"  {f}")
    else:
        print("\nエラー: 生成できるレポートがありません。input/フォルダにファイルを配置してください。")
        sys.exit(1)


if __name__ == "__main__":
    main()
